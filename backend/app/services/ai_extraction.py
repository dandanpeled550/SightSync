"""
AI extraction service: parse an xlsx file and call an AI provider to extract tasks.

Provider selection (checked in order):
  1. ANTHROPIC_API_KEY set → use Claude
  2. OPENAI_API_KEY set    → use OpenAI
  3. Neither set           → return error, never raise

Never raises exceptions — all failures are captured in ExtractionResult.error.
"""

import io
import json
import logging
import time
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl

from app.config import settings

logger = logging.getLogger(__name__)


@dataclass
class ExtractedTask:
    name: str
    level_tag: str
    trade_tag: Optional[str]
    start_date: str   # ISO date string "YYYY-MM-DD"
    duration_days: int
    apartment_tag: Optional[str]
    room_tag: Optional[str]
    excel_row_index: int = 0
    workflow_id: str = ""     # assigned by Pass 2; empty string if Pass 2 skipped


@dataclass
class Workflow:
    """A named group of tasks belonging to the same trade, ordered by level."""
    id: str                   # "wf_0", "wf_1", ...
    name: str                 # e.g. "Electrical", "Structural"
    task_indices: list        # ordered task indices within this workflow


@dataclass
class InferredDependency:
    """A single dependency edge inferred by Pass 2."""
    task_index: int
    depends_on_index: int
    lag_days: int
    confidence: float
    reasoning: str
    type: str                 # "intra_workflow" | "cross_workflow_handoff"


@dataclass
class ExtractionResult:
    tasks: list = field(default_factory=list)
    workflows: list = field(default_factory=list)
    dependencies: list = field(default_factory=list)
    confidence: float = 0.0
    error: Optional[str] = None
    raw_text_length: int = 0


# Minimal valid styles.xml — replaces invalid styles to bypass openpyxl's colour-pattern
# validation (MatchPattern raises "The string did not match the expected pattern." on
# non-standard ARGB values like 3-char hex or vendor-specific colour codes).
_MINIMAL_STYLES_XML = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
    '<fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>'
    '<fills count="2">'
    '<fill><patternFill patternType="none"/></fill>'
    '<fill><patternFill patternType="gray125"/></fill>'
    '</fills>'
    '<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>'
    '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
    '<cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>'
    '</styleSheet>'
)


def _strip_xlsx_styles(xlsx_bytes: bytes) -> bytes:
    """Replace xl/styles.xml with a minimal valid stub so cell values can be read."""
    buf_in = io.BytesIO(xlsx_bytes)
    buf_out = io.BytesIO()
    with zipfile.ZipFile(buf_in, 'r') as zin, \
         zipfile.ZipFile(buf_out, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = _MINIMAL_STYLES_XML.encode() if item.filename == 'xl/styles.xml' \
                   else zin.read(item.filename)
            zout.writestr(item, data)
    return buf_out.getvalue()


_TASK_SHEET_KEYWORDS = {"task", "schedule", "gantt", "work plan", "workplan", "activity", "activities"}


def _sheet_priority(sheet_title: str) -> int:
    """Lower number = appears first. Task-named sheets get priority 0."""
    lower = sheet_title.lower()
    return 0 if any(kw in lower for kw in _TASK_SHEET_KEYWORDS) else 1


def _parse_workbook(wb) -> list:
    """Extract non-empty rows from all sheets as tab-separated strings.

    Task-named sheets are emitted first so they are never truncated out
    when many irrelevant sheets precede them in the workbook.
    """
    out = []
    sheets = sorted(wb.worksheets, key=lambda s: _sheet_priority(s.title))
    for sheet in sheets:
        out.append(f"Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            try:
                row_values = [str(cell) if cell is not None else "" for cell in row]
            except Exception:
                continue
            if any(v.strip() for v in row_values):
                out.append("\t".join(row_values))
    wb.close()
    return out


def _xlsx_to_text(xlsx_bytes: bytes) -> tuple:
    """Parse xlsx bytes into flat text (max 20000 chars). Returns (text, length)."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
        lines = _parse_workbook(wb)
    except Exception:
        # Fallback: strip xl/styles.xml to bypass colour/style pattern errors.
        # Cell values live in sheet XMLs — styles only affect formatting, not data.
        cleaned = _strip_xlsx_styles(xlsx_bytes)
        wb = openpyxl.load_workbook(io.BytesIO(cleaned), data_only=True)
        lines = _parse_workbook(wb)

    full_text = "\n".join(lines)
    truncated = full_text[:20000]
    return truncated, len(truncated)


def _load_prompt(name: str) -> str:
    """Load backend/app/services/prompts/{name}.md relative to this file."""
    prompt_dir = Path(__file__).parent / "prompts"
    return (prompt_dir / f"{name}.md").read_text()


def _build_prompt(text: str) -> str:
    template = _load_prompt("task_extraction")
    return template.replace("{{XLSX_TEXT}}", text)


def _call_anthropic(prompt: str) -> str:
    """Call Claude and return the raw response text. Raises on any error."""
    import anthropic
    logger.debug("AI prompt [anthropic/%s] (%d chars):\n%s", settings.anthropic_model, len(prompt), prompt)
    start = time.perf_counter()
    client = anthropic.Anthropic(api_key=settings.anthropic_api_key)
    message = client.messages.create(
        model=settings.anthropic_model,
        max_tokens=4096,
        messages=[{"role": "user", "content": prompt}],
    )
    text = message.content[0].text
    elapsed = int((time.perf_counter() - start) * 1000)
    logger.info("AI call [anthropic/%s] done in %dms — %d chars", settings.anthropic_model, elapsed, len(text))
    logger.debug("AI response:\n%s", text)
    return text


def _call_openai(prompt: str) -> str:
    """Call OpenAI and return the raw response text. Raises on any error."""
    import openai
    logger.debug("AI prompt [openai/%s] (%d chars):\n%s", settings.openai_model, len(prompt), prompt)
    start = time.perf_counter()
    client = openai.OpenAI(api_key=settings.openai_api_key)
    # No response_format param — SDK version differences cause Pydantic validation errors.
    # _strip_fences() handles markdown-wrapped responses as a fallback.
    response = client.chat.completions.create(
        model=settings.openai_model,
        max_tokens=4096,
        messages=[{"role": "user", "content": prompt}],
    )
    text = response.choices[0].message.content
    elapsed = int((time.perf_counter() - start) * 1000)
    logger.info("AI call [openai/%s] done in %dms — %d chars", settings.openai_model, elapsed, len(text))
    logger.debug("AI response:\n%s", text)
    return text


def _strip_fences(text: str) -> str:
    """Strip markdown code fences if present (e.g. ```json ... ```)."""
    text = text.strip()
    if text.startswith("```"):
        lines = text.splitlines()
        # Remove first line (```json or ```) and last line (```)
        inner = lines[1:] if len(lines) > 1 else lines
        if inner and inner[-1].strip() == "```":
            inner = inner[:-1]
        text = "\n".join(inner).strip()
    return text


def _parse_response(response_text: str, text_length: int) -> ExtractionResult:
    """Parse the JSON response from any provider into an ExtractionResult."""
    data = json.loads(_strip_fences(response_text))
    raw_tasks = data.get("tasks", [])
    confidence = float(data.get("confidence", 0.0))

    tasks = []
    for i, t in enumerate(raw_tasks):
        tasks.append(
            ExtractedTask(
                name=str(t["name"]),
                level_tag=str(t["level_tag"]),
                trade_tag=t.get("trade_tag") or None,
                start_date=str(t["start_date"]),
                duration_days=int(t["duration_days"]),
                apartment_tag=t.get("apartment_tag") or None,
                room_tag=t.get("room_tag") or None,
                excel_row_index=i,
                workflow_id="",
            )
        )
    return ExtractionResult(
        tasks=tasks,
        workflows=[],
        dependencies=[],
        confidence=confidence,
        error=None,
        raw_text_length=text_length,
    )


def infer_workflows_and_dependencies(
    tasks,
    settings_obj=None,
) -> tuple:
    """Pass 2: inject tasks JSON into dependency_inference.md → call Claude.

    Filter deps where confidence < 0.4.
    On any failure: return ([], []) — never raise.
    """
    if settings_obj is None:
        settings_obj = settings

    try:
        # Build tasks JSON for the prompt
        tasks_data = []
        for i, t in enumerate(tasks):
            tasks_data.append({
                "index": i,
                "name": t.name,
                "level_tag": t.level_tag,
                "trade_tag": t.trade_tag,
                "start_date": t.start_date,
                "duration_days": t.duration_days,
            })
        tasks_json = json.dumps(tasks_data, indent=2)

        template = _load_prompt("dependency_inference")
        prompt = template.replace("{{TASKS_JSON}}", tasks_json)

        # Call Claude for Pass 2 (only Anthropic supported for Pass 2)
        import anthropic
        client = anthropic.Anthropic(api_key=settings_obj.anthropic_api_key)
        message = client.messages.create(
            model=settings_obj.anthropic_model,
            max_tokens=4096,
            messages=[{"role": "user", "content": prompt}],
        )
        response_text = message.content[0].text

        data = json.loads(_strip_fences(response_text))

        # Parse workflows
        workflows = []
        for wf_data in data.get("workflows", []):
            workflows.append(Workflow(
                id=str(wf_data["id"]),
                name=str(wf_data["name"]),
                task_indices=list(wf_data.get("task_indices", [])),
            ))

        # Parse dependencies — filter out low-confidence ones
        dependencies = []
        for dep_data in data.get("dependencies", []):
            confidence = float(dep_data.get("confidence", 0.0))
            if confidence < 0.4:
                continue
            dependencies.append(InferredDependency(
                task_index=int(dep_data["task_index"]),
                depends_on_index=int(dep_data["depends_on_index"]),
                lag_days=int(dep_data.get("lag_days", 0)),
                confidence=confidence,
                reasoning=str(dep_data.get("reasoning", "")),
                type=str(dep_data.get("type", "intra_workflow")),
            ))

        return workflows, dependencies

    except Exception as exc:
        logger.warning("Pass 2 (dependency inference) failed: %s", exc)
        return [], []


def extract_tasks_from_xlsx(xlsx_bytes: bytes) -> ExtractionResult:
    """
    Parse xlsx bytes, call the available AI provider, and return an ExtractionResult.

    Provider priority: Anthropic (Claude) > OpenAI (GPT).
    Never raises — all errors are captured in ExtractionResult.error.

    Pass 1: Extract tasks from xlsx text.
    Pass 2: Infer workflow groupings and dependencies (Anthropic only; gracefully
            degrades to empty lists if it fails or no Anthropic key is set).
    """
    # Detect provider
    use_anthropic = bool(settings.anthropic_api_key)
    use_openai = bool(settings.openai_api_key)

    if not use_anthropic and not use_openai:
        return ExtractionResult(
            tasks=[],
            confidence=0.0,
            error="No AI API key configured — set ANTHROPIC_API_KEY or OPENAI_API_KEY",
        )

    # Parse xlsx to text
    try:
        text, text_length = _xlsx_to_text(xlsx_bytes)
    except Exception as e:
        return ExtractionResult(tasks=[], confidence=0.0, error=str(e))

    # Call provider (Claude takes priority if both keys are present)
    prompt = _build_prompt(text)
    try:
        if use_anthropic:
            response_text = _call_anthropic(prompt)
        else:
            response_text = _call_openai(prompt)
    except Exception as e:
        return ExtractionResult(tasks=[], confidence=0.0, error=str(e), raw_text_length=text_length)

    # Parse JSON response (Pass 1)
    try:
        result = _parse_response(response_text, text_length)
    except Exception as e:
        return ExtractionResult(tasks=[], confidence=0.0, error=str(e), raw_text_length=text_length)

    # Pass 2: infer workflows and dependencies (Anthropic only)
    # Only run if we have tasks and an Anthropic key
    pass2_error = None
    if result.tasks and use_anthropic:
        try:
            workflows, dependencies = infer_workflows_and_dependencies(result.tasks, settings)
            result.workflows = workflows
            result.dependencies = dependencies

            # Assign workflow_id to each task based on which workflow it belongs to
            for workflow in workflows:
                for task_idx in workflow.task_indices:
                    if 0 <= task_idx < len(result.tasks):
                        result.tasks[task_idx].workflow_id = workflow.id
        except Exception as exc:
            logger.warning("Pass 2 assignment failed: %s", exc)
            pass2_error = str(exc)
            result.workflows = []
            result.dependencies = []

    # If Pass 2 returned empty (failed internally), set error note if not already set
    if result.tasks and use_anthropic and not result.workflows and not result.dependencies and pass2_error:
        result.error = f"Pass 2 failed: {pass2_error}"

    return result
